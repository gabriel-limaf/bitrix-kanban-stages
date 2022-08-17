"""
Codigo responsavel pela exportacao dos stages_id e stage_title das etapas do kanban de cada workgroup
"""

import requests
import pandas as pd
from openpyxl import load_workbook
from time import sleep
import json


base_url = ""
task_url = "task.stages.get"
url = base_url + task_url

path_saida = ''
df1 = pd.DataFrame(pd.read_excel(path_saida, sheet_name='Projetos'))


for i, row in df1.iterrows():
    workgroupId = str(row['Workgroup_ID'])
    payload = json.dumps(
        {"entityId": workgroupId})
    headers = {
        'Content-Type': 'application/json',
    }
    response = requests.request("POST", url, headers=headers, data=payload)
    print(response.text)
    obj = json.loads(response.text)
    sleep(1)
    if response.status_code != 200:
        idSaida = (obj['error'])
        df1.loc[i, 'status_api'] = idSaida
        book = load_workbook(path_saida)
        writer = pd.ExcelWriter(path_saida, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df1.to_excel(writer, "Projetos", header=True, index=False)
        writer.save()
        sleep(1)
        continue
    else:
        for item in obj['result'].values():
            stageId = (item['ID'])
            stageTitle = (item['TITLE'])
            print(stageId)
            print(stageTitle)
            if stageTitle in df1.columns:
                df1.loc[i, stageTitle] = stageId
                book = load_workbook(path_saida)
                writer = pd.ExcelWriter(path_saida, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df1.to_excel(writer, "Projetos", header=True, index=False)
                writer.save()
